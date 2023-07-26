import pandas as pd
import os
from openpyxl import load_workbook

DATA_DIR = "data/transformed"
os.makedirs(DATA_DIR, exist_ok=True)

def read_data(**args):
    #unpack args.
    filepath = args['filepath']
    sheet_name = args['sheet_name']
    skiprows = args['skiprows']
    nrows = args['nrows']
    engine = args['engine']
    data = pd.read_excel(filepath, sheet_name=sheet_name, skiprows=skiprows, nrows=nrows, engine=engine)

    #remove repeat colum headings
    data = data[data.Activity != "Activity"]
    return data

def comments_to_csv(data, filepath, sheet_name, filename, colname, min_col, max_col):
    #only take filled cells
    data = data[data.loc[:, colname].notnull()]
    #load the data
    workbook = load_workbook(filepath)
    worksheet = workbook[sheet_name]
    #create empty list
    description = []

    #get the comment for each relevant cell in the "type" column.
    for row in worksheet.iter_rows(min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.comment:
                comm = cell.comment.text
                arr = comm.split("Comment:")
                description.append(arr[1])

    transport_type = data.loc[:, colname].to_list()
    assert len(transport_type) ==  len(description)
    
    info = dict()
    for i, j in zip(transport_type, description):
        info.update({i:j})

    info_as_frame = pd.Series(data=info, name='Description')
    info_as_frame = info_as_frame.replace("\n", "")
    info_as_frame = info_as_frame.str.strip()
    info_as_frame.to_csv(os.path.join(DATA_DIR, filename))
    return info_as_frame

def transform_data(data, use_cols, mapper=None):
    data = data.loc[:, use_cols]
    if mapper:
        data.rename(columns=mapper, inplace=True)
    return data

if __name__ == "__main__":
    fp = 'data-raw/gov_emissions.xlsx'
    engine = "openpyxl"

    passenger_vehicles = read_data(filepath=fp, sheet_name="Passenger vehicles", skiprows=24, nrows=43, engine=engine)
    pv_data = transform_data(passenger_vehicles, use_cols=["Activity", "Type", "Unit", "kg CO2e", "kg CO2e.1", "kg CO2e.2", "kg CO2e.3"],
                             mapper={"kg CO2e": "Diesel", "kg CO2e.1": "Petrol", "kg CO2e.2": "Unknown", "kg CO2e.3": "Plugin hybrid electric vehicle"})
    pv_data.to_csv(os.path.join(DATA_DIR, 'passenger_vehicles.csv'))
    comments_to_csv(pv_data, filepath='data-raw/gov_emissions.xlsx', 
                    sheet_name="Passenger vehicles", 
                    filename='passenger_vehicles_info.csv',
                    colname='Type', 
                    min_col=1, 
                    max_col=2)

    air_travel = read_data(filepath=fp, sheet_name="Business travel- air", skiprows=21, nrows=15, engine=engine)
    air_travel = transform_data(air_travel, use_cols=['Activity', 'Haul', 'Class', 'Unit', 'kg CO2e'], mapper={'Haul': 'Type'})
    air_travel.to_csv(os.path.join(DATA_DIR, 'air_travel.csv'))
    comments_to_csv(air_travel, 
                    filepath='data-raw/gov_emissions.xlsx', 
                    sheet_name="Business travel- air", 
                    filename='air_travel_info.csv',
                    colname='Type', 
                    min_col=0, 
                    max_col=2)
    
    sea_travel = read_data(filepath=fp, sheet_name="Business travel- sea", skiprows=16, nrows=5, engine=engine)
    sea_travel['Activity'] = 'Ferry'
    sea_travel.to_csv(os.path.join(DATA_DIR, 'sea_travel.csv'))

    taxis = read_data(filepath=fp, sheet_name="Business travel- land", skiprows=69, nrows=5, engine=engine)
    buses = read_data(filepath=fp, sheet_name="Business travel- land", skiprows=77, nrows=5, engine=engine)
    trains = read_data(filepath=fp, sheet_name="Business travel- land", skiprows=85, nrows=5, engine=engine)

    taxis = transform_data(taxis, use_cols=["Activity", "Type", "Unit", "kg CO2e"])
    taxis['Activity'] = 'Taxis'
    buses = transform_data(buses, use_cols=["Activity", "Type", "Unit", "kg CO2e"])
    buses['Activity'] = 'Bus'
    trains = transform_data(trains, use_cols=["Activity", "Type", "Unit", "kg CO2e"])
    trains['Activity'] = 'Rail'

    taxis.to_csv(os.path.join(DATA_DIR, 'taxis.csv'))
    buses.to_csv(os.path.join(DATA_DIR, 'buses.csv'))
    trains.to_csv(os.path.join(DATA_DIR, 'trains.csv'))